"""Convert gold_items.jsonl to a user-friendly Excel file for annotation.

Also inserts trap items (flipped system_score) to detect lazy/AI annotation.
Trap mapping is saved separately for verification.

Usage:
    python calibration/export_to_excel.py
"""

import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INPUT = Path("calibration/gold_items.jsonl")
OUTPUT = Path("calibration/标注任务.xlsx")
TRAP_MAP = Path("calibration/.trap_mapping.json")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

random.seed(42)


def load_items():
    items = []
    with open(INPUT, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                items.append(json.loads(line))
    return items


def insert_traps(items: list[dict]) -> list[dict]:
    """Flip system_score on ~8 items to create detectable traps."""
    trap_candidates = []
    for i, item in enumerate(items):
        if item["item_type"] == "binary":
            trap_candidates.append(("binary", i))
        elif item["item_type"] == "dimension":
            score = item.get("system_score")
            if isinstance(score, (int, float)) and score >= 3:
                trap_candidates.append(("dimension", i))

    random.shuffle(trap_candidates)
    traps = trap_candidates[:8]

    trap_map = []
    for item_type, idx in traps:
        original = items[idx]["system_score"]
        if item_type == "binary":
            items[idx]["system_score"] = not original
            trap_map.append(
                {
                    "index": idx,
                    "item_id": items[idx]["item_id"],
                    "scenario": items[idx]["scenario_name"],
                    "original": original,
                    "flipped_to": items[idx]["system_score"],
                }
            )
        else:
            flipped = max(0, original - 3) if original >= 3 else min(5, original + 3)
            items[idx]["system_score"] = flipped
            trap_map.append(
                {
                    "index": idx,
                    "item_id": items[idx]["item_id"],
                    "scenario": items[idx]["scenario_name"],
                    "original": original,
                    "flipped_to": flipped,
                }
            )

    return items, trap_map


def build_excel(items: list[dict]):
    wb = Workbook()

    # -- 说明页 --
    ws_info = wb.active
    ws_info.title = "说明"
    instructions = [
        ["标注任务说明"],
        [""],
        ["1. 切到「维度评分」sheet，在「你的评分」列填 0-5 的整数"],
        ["2. 切到「安全检查」sheet，在「你的判断」列填 TRUE 或 FALSE"],
        ["3. 「系统评分」列是系统给的参考分，你可以参考但不一定要同意"],
        ["4. 如果信息不够判断，维度评分填 -1，安全检查留空"],
        ["5. 「备注」列可选，写你不同意系统分数的理由"],
        ["6. 做完后把这个文件发回来"],
        [""],
        ["预计耗时：1-2 小时"],
        ["不需要全部做完，做 60 条以上就有统计意义"],
        ["建议先做「安全检查」（简单），再做「维度评分」"],
    ]
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=row[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        else:
            cell.font = Font(size=11)
    ws_info.column_dimensions["A"].width = 80

    dims = [i for i in items if i["item_type"] == "dimension"]
    bins = [i for i in items if i["item_type"] == "binary"]

    # -- 维度评分页 --
    ws_dim = wb.create_sheet("维度评分")
    dim_headers = [
        "序号",
        "场景",
        "难度",
        "维度",
        "评分标准",
        "对话摘要",
        "系统评分",
        "你的评分",
        "备注",
    ]
    for col, h in enumerate(dim_headers, 1):
        cell = ws_dim.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    for row_idx, item in enumerate(dims, 2):
        conv = item.get("conversation_summary", "")
        if len(conv) > 500:
            conv = conv[:500] + "..."
        values = [
            row_idx - 1,
            item["scenario_name"],
            item["difficulty"],
            item["item_name"],
            item["rubric"],
            conv,
            item["system_score"],
            None,
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws_dim.cell(row=row_idx, column=col, value=val)
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    dim_widths = [5, 25, 8, 15, 50, 60, 8, 8, 30]
    for i, w in enumerate(dim_widths, 1):
        ws_dim.column_dimensions[get_column_letter(i)].width = w
    ws_dim.freeze_panes = "A2"

    # -- 安全检查页 --
    ws_bin = wb.create_sheet("安全检查")
    bin_headers = [
        "序号",
        "场景",
        "难度",
        "检查项",
        "判断标准",
        "对话摘要",
        "系统判断",
        "你的判断",
        "备注",
    ]
    for col, h in enumerate(bin_headers, 1):
        cell = ws_bin.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    for row_idx, item in enumerate(bins, 2):
        conv = item.get("conversation_summary", "")
        if len(conv) > 500:
            conv = conv[:500] + "..."
        sys_label = "是" if item["system_score"] else "否"
        values = [
            row_idx - 1,
            item["scenario_name"],
            item["difficulty"],
            item["item_name"],
            item["rubric"],
            conv,
            sys_label,
            None,
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws_bin.cell(row=row_idx, column=col, value=val)
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    bin_widths = [5, 25, 8, 15, 50, 60, 8, 8, 30]
    for i, w in enumerate(bin_widths, 1):
        ws_bin.column_dimensions[get_column_letter(i)].width = w
    ws_bin.freeze_panes = "A2"

    wb.save(OUTPUT)
    return len(dims), len(bins)


def main():
    items = load_items()
    items, trap_map = insert_traps(items)
    dim_count, bin_count = build_excel(items)

    with open(TRAP_MAP, "w", encoding="utf-8") as f:
        json.dump(trap_map, f, ensure_ascii=False, indent=2)

    print(f"已生成: {OUTPUT}")
    print(f"  维度评分: {dim_count} 条")
    print(f"  安全检查: {bin_count} 条")
    print(f"  陷阱项: {len(trap_map)} 条（映射保存在 {TRAP_MAP}）")
    print()
    print("陷阱项详情（不要让队友看到）：")
    for t in trap_map:
        print(f"  [{t['item_id']}] {t['scenario']} — 原值 {t['original']} → 改成 {t['flipped_to']}")


if __name__ == "__main__":
    main()
